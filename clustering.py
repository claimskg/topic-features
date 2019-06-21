import sys, getopt
import pickle
import pandas as pd
import numpy as np
import collections
import nltk

from rdflib import Graph
from nltk.corpus import wordnet
from nltk.stem import WordNetLemmatizer
from matplotlib import pyplot as plt
from sklearn.preprocessing import normalize

from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
from sklearn.cluster import KMeans

from openpyxl import Workbook

def get_keyword_provenance(sorted_dict,df):
    organization_dict = dict.fromkeys(sorted_dict)
    for index, row in df.iterrows():
        keywords = [x.strip() for x in row.keywords.split(',')]
        author = row.organization
        for keyword in keywords:
            if keyword in organization_dict:
                if organization_dict[keyword] == None:
                    organization_dict[keyword] = author
                else:
                    if organization_dict[keyword].find(author) == -1:
                        organization_dict[keyword] += ', ' + author
    return organization_dict
	
def get_clusters(labels,keywordList):
    clusters = dict()
    for wordIndex, clusterIndex in enumerate(labels):
        if clusterIndex > -1:
            if clusterIndex in clusters:
                clusters[clusterIndex] += ", " + keywordList[wordIndex]
            else:
                clusters[clusterIndex] = keywordList[wordIndex]
    return clusters

def get_cluster_distribution(clusters,df):
    cluster_dist = dict()
    for x in range(0, len(clusters)):
        cluster_dist[x] = dict()
        cluster_dist[x]["TRUE"] = 0
        cluster_dist[x]["FALSE"] = 0
        cluster_dist[x]["MIXTURE"] = 0
        cluster_dist[x]["OTHER"] = 0
        cluster_dist[x]["TOTAL"] = 0

    for index, row in df.iterrows():
        keywords = [x.strip() for x in row.keywords.split(',')]
        rating = row.rating
        for x in range(0, len(clusters)):
            kwl = [y.strip() for y in clusters[x].split(',')]
            for keyword in keywords: 
                if keyword in kwl:
                    cluster_dist[x]["TOTAL"] += 1
                    cluster_dist[x][rating] += 1
                    break
    return cluster_dist
        
def add_cluster_to_excel(wb,name,clusters,clust_dist):
    ws = wb.create_sheet(name)
    ws.append(["Cluster", "#Claims", "#True","#False","#Mixture","#Other","Keywords"])
    for x in range(0, len(clusters)):
        row = [x+1,clust_dist[x]["TOTAL"],clust_dist[x]["TRUE"],clust_dist[x]["FALSE"]
              ,clust_dist[x]["MIXTURE"],clust_dist[x]["OTHER"]]
        row.extend(clusters[x].split(','))
        ws.append(row)

def main(argv):
	#Parameters
	csvfile = 'Data/data.csv'
	keywordsFile = "Data/Keywords.xlsx"
	n_occurrences = 50
	wardFile = "Data/Ward.xlsx"
	start = 10
	step = 10
	stop = 30
	
	try:
		opts, args = getopt.getopt(argv,"i:c:k:n:r:")
	except getopt.GetoptError:
		print('clustering.py -i <input_csv> -c <output_clustering> -k <output_keywords> -n <keyword_occurrence_floor> -r <start,stop,step>')
		sys.exit(2)
	
	for opt, arg in opts:
		if opt == '-h':
			print('clustering.py -i <input_csv> -c <output_clustering> -k <output_keywords> -n <keyword_occurrence_floor>')
			sys.exit()
		elif opt in ("-i"):
			csvfile = arg
		elif opt in ("-c"):
			wardFile = arg
		elif opt in ("-k"):
			keywordsFile = arg
		elif opt in ("-n"):
			try:
				n_occurrences = int(arg)
			except ValueError:
				n_occurrences = 50
		elif opt in ("-r"):
			ra = [x.strip() for x in arg.split(",")]
			if(len(ra) == 3):
				try:
					start = int(ra[0])
				except ValueError:
					start = 10
				try:
					stop = int(ra[1])
				except ValueError:
					stop = 30
				try:
					step = int(ra[2])
				except ValueError:
					step = 10
	
	stop += step
	
	#Read csv
	df = pd.read_csv(csvfile,sep=";")
	
	#Get vocabulary
	keywordDict = dict()
	for index, row in df.iterrows():
		keywords = [x.strip() for x in row.keywords.split(',')]
		for keyword in keywords:
			if keyword in keywordDict:
				keywordDict[keyword] += 1
			else:
				keywordDict[keyword] = 1
				
	sorted_dict = sorted(keywordDict.items(), key=lambda kv: kv[1])
	sorted_dict = collections.OrderedDict(sorted_dict)

	#Keep keywords with n or more ocurrences
	for key in list(sorted_dict):
		if sorted_dict[key] < n_occurrences:
			del sorted_dict[key]

	keywordList = list(sorted_dict)
	
	organization_dict = get_keyword_provenance(sorted_dict,df)
	
	#Save keywords to excel
	wb = Workbook()
	sheet1 = wb['Sheet']
	wb.remove(sheet1)

	ws = wb.create_sheet("Keywords")
	ws.append(["Keyword", "Occurrences", "Sites"])

	for key in sorted_dict:
		row = [key,sorted_dict[key],organization_dict[key]]
		ws.append(row)
		
	wb.save(keywordsFile)
	
	#Build co occurrence matrix
	n = len(keywordList)
	matrix = np.zeros((n,n))

	for index, row in df.iterrows():
		keywords = [x.strip() for x in row.keywords.split(',')]
		for k1 in keywords:
			try:
				i1 = keywordList.index(k1)
			except ValueError:
				continue
			for k2 in keywords:
				if k1 != k2:
					try:
						i2 = keywordList.index(k2)
					except ValueError:
						continue
					matrix[i1][i2] = matrix[i1][i2]+1
	
	#Cosine similarity
	cs_matrix = cosine_similarity(matrix, matrix)
	
	#Ward Clustering
	wb = Workbook()
	sheet1 = wb['Sheet']
	wb.remove(sheet1)

	#Ward with Cosine Similarity Matrix
	for n in range(start,stop,step):
		ward = AgglomerativeClustering(n_clusters=n).fit(cs_matrix)
		clusters = get_clusters(ward.labels_,keywordList)
		clust_dist = get_cluster_distribution(clusters,df)
		add_cluster_to_excel(wb,str(n) + " Clusters",clusters,clust_dist)

	#Save Excel File
	wb.save(wardFile)
	
   
if __name__ == "__main__":
	main(sys.argv[1:])